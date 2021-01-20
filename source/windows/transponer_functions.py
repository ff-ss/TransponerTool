

import consolemenu
from colorama import init
from colorama import Fore, Back, Style
import openpyxl
import os
import time
from pathlib import Path


# DICTIONARY DECLARATION
columnDictionary = {
    "A" : 1,
    "B" : 2,
    "C" : 3,
    "D" : 4,
    "E" : 5,
    "F" : 6,
    "G" : 7,
    "H" : 8,
    "I" : 9,
    "J" : 10,
    "K" : 11,
    "L" : 12,
    "M" : 13,
    "N" : 14,
    "O" : 15,
    "P" : 16,
    "Q" : 17,
    "R" : 18,
    "S" : 19,
    "T" : 20,
    "U" : 21,
    "V" : 22,
    "W" : 23,
    "X" : 24,
    "Y" : 25,
    "Z" : 26,
    "NONE":-1

}

# COLOR FUNCTIONS

def outputcolor():
    print(Fore.WHITE + Back.MAGENTA + Style.BRIGHT)
def inputcolor():
    print(Fore.GREEN + Back.BLACK + Style.BRIGHT)
def resetandclose():
    print(Style.RESET_ALL)

#VARIBALE DEFINITION

menu = None
wd = None
selection = None

filename = None
file = None
o = None
output = None

sheet_old = None
sheet_new = None

scolumnname = None
idcolumn_old = None
idcolumn_new = None
startcolumn = None
endcolumn = None


idrow_old = None
idrow_new = 1
cells = 0
TAG = None
filter = None


# ERROR HANDLING
def trycolumn(checkFor = "default"):
    if checkFor == "default":
        try:
            columnDictionary[scolumnname]
        except:
            outputcolor()
            print("                        ERROR:                            ")
            print("      THE EXCEL COLUMN YOU HAVE SELECTED IS INVALID.       ")
            print("                      TRY AGAIN                            ")
            return False
        else:
            return True
    else:
        return True


def mainMenu():
    menu = consolemenu.SelectionMenu(["START"],title="TrasponerTool", subtitle="USE THIS TOOL TO EASILY TRANSPOSE EXCEL CELLS",prologue_text="THIS PROGRAM USES THE MODULES OPENPYXL, COLORAMA, OS, TIME, AND CONSOLEMENU", epilogue_text="ABOUT: Version 3.2.0 (C) 2021 github.com/ff-ss",show_exit_option=False)
    menu.show()
    menu.join()


def columnWarning():
    print("______________________________________________________________________________")
    print("|             YOU WILL START BY DECLARING YOUR FILE'S VARIABLES               |")
    print("|         PLEASE, ENTER ONLY THE LETTER INDICATING THE EXCEL COLUMN           |")
    print("|                                                                             |")
    print("|     PLEASE REMEMBER, THE PROGRAM WILL STOP AT THE FIRST FULLY EMPTY ROW     |")
    print("|                              USE CAREFULLY                                  |")
    print("|                                                                             |")
    print("|                     PRESS THE RETURN KEY TO CONTINUE                        |")
    print("|_____________________________________________________________________________|")
    input("\n>> ")
    os.system("cls")
# INPUT FUNCTIONS

def newquery():
    global exitcode
    options= ["CONTINUE WITH THIS FILE (ANOTHER SHEET)", "EXPORT"]
    menu = consolemenu.SelectionMenu(options,title="DO YOU WISH TO CONTINUE USING THIS FILE?",subtitle="EXPORT TO PRODUCE A SINGLE FILE",epilogue_text=" ",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    if selection == 0:
        return True
    elif selection == 1:
        return False

def setDirectory():

    cwd = os.getcwd()
    subwd = next(os.walk('.'))[1]
    subwd.append('THIS IS MY FOLDER')
    if not subwd:
        print("___________________________________________________________")
        print("|                                                         |")
        print("|             THERE ARE NO MORE SUBFOLDERS HERE           |")
        print("|            SETTING THIS FOLDER AS THE DIRECTORY...      |")
        print("|          PLEASE RESTART IF THIS NOT YOUR FOLDER         |")
        print("|_________________________________________________________|")
        time.sleep(2)
        return cwd
    menu = consolemenu.SelectionMenu(subwd,title="LET'S LOOK UP THE FOLDER WHERE YOUR FILE IS",subtitle=f"WE ARE CURRENTLY IN {cwd}",epilogue_text="YOU CAN'T GO BACK, PLEASE RESTART IF YOU FAILED TO CHOOSE A CORRECT PATH",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    print(selection)
    sub = subwd[selection]
    print(sub)
    if sub == 'THIS IS MY FOLDER':
        cwd = os.getcwd()
        return cwd
    else:
        os.chdir(cwd + "/"+sub)

    cwd = os.getcwd()
    print(cwd)
    return setDirectory()





def openFile(path):
    global wd
    global menu
    global selection
    global filename
    global file
    global o

    os.chdir(path)

    wd = os.listdir()
    print(wd)
    menu = consolemenu.SelectionMenu(wd,title="PLEASE SELECT THE SPREADSHEET FROM WHERE TO EXTRACT DATA",subtitle="USE ONLY EXCEL (.XLS .XLSX) FILES",epilogue_text="IF YOU CAN'T FIND YOUR FILE, MAKE SURE YOU ARE RUNNING THIS PROGRAM ON THE SAME FOLDER AS YOUR FILES.",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    filename = wd[selection]
    try:
        openpyxl.load_workbook(filename)
    except:
        print("                                 ERROR:                                   ")
        print("                        YOU HAVE PICKED AN INVALID FILE                   ")
        print("                                 TRY AGAIN                                ")
        input(">> ")
        openFile()
    else:
        file = openpyxl.load_workbook(filename)
        o = openpyxl.Workbook()
        time.sleep(0.2)

def openSheet():
    global wd
    global menu
    global selection
    global sheet_old
    wd = file.sheetnames
    menu = consolemenu.SelectionMenu(wd,title="CHOOSE THE SHEET FROM WHERE TO EXTRACT DATA.",subtitle="PLEASE CHECK THIS IS RIGHT SHEET",epilogue_text="YOU MAY OPEN YOUR FILE TO VERIFY",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    global fixdopt
    global fixddir
    fixdopt = menu.selected_option
    fixddir = file.sheetnames
    sheet_old = file.worksheets[file.sheetnames.index(wd[selection])]

    time.sleep(0.2)

def assignColumns():

    global scolumnname
    global idcolumn_old
    global startcolumn
    global endcolumn
    global TAG
    global filter

    print("___________________________________________________________")
    print("|                                                         |")
    print("|             ENTER THE COLUMN LETTER WHERE               |")
    print("|                   YOUR KEY IS FOUND                     |")
    print("|                   EJ. IDS, NAMES, ...                   |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    idcolumn_old = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")


    print("___________________________________________________________")
    print("|                                                         |")
    print("| ENTER THE COLUMN LETTER WHERE YOUR VALUE RANGE STARTS:  |")
    print("|                   EJ. ITEM, DATA...                     |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    startcolumn = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")

    print("___________________________________________________________")
    print("|                                                         |")
    print("|   ENTER THE COLUMN LETTER WHERE YOUR VALUE RANGE ENDS:  |")
    print("|                   EJ. ITEM, DATA...                     |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    endcolumn = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")

    print("___________________________________________________________")
    print("|                                                         |")
    print("|   ENTER THE COLUMN LETTER WHERE YOUR FILTER COLUMN IS,  |")
    print("|       OR PRESS RETURN TO SKIP THE FILTER COLUMN         |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    if scolumnname == "":
        stat = "nofilter"
    else:
        stat = "default"
    while not trycolumn(stat):
        scolumnname = input(">> ").capitalize()
    if scolumnname == "":
        filter = columnDictionary["NONE"]
    else:
        filter = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")

    print("___________________________________________________________")
    print("|                                                         |")
    print("|                                                         |")
    print("|      CREATE A TAG, TO IDENTIFY CELLS FROM THE SHEET     |")
    print(f"|                   {fixddir[fixdopt]}                    |")
    print("|_________________________________________________________|")
    TAG = input("\n>> ")
    time.sleep(0.2)
    os.system("cls")

def setExport():
    global sheet_new
    global output
    print("___________________________________________________________")
    print("|                                                         |")
    print("|              NAME YOUR EXPORT EXCEL FILE                |")
    print("|     USE A NEW NAME TO AVOID OVERWRITING OTHERS          |")
    print("|_________________________________________________________|")

    output = input("\n>> ")+"export.xlsx"
    sheet_new = o.worksheets[0]
    time.sleep(0.2)
    os.system("cls")


def compute():

    global idrow_new
    global idrow_old
    global exitcode
    global cells

    idrow_old = 1
    idcolumn_new = 1

    while sheet_old.cell(idrow_old, idcolumn_old).value is not None:
        for column in range(startcolumn,endcolumn+1):
            if sheet_old.cell(idrow_old, column).value is not None:
                sheet_new.cell(idrow_new, idcolumn_new).value = sheet_old.cell(idrow_old, idcolumn_old).value

                if sheet_old.cell(idrow_old, column).value is not None:
                    sheet_new.cell(idrow_new,2).value = sheet_old.cell(idrow_old, column).value
                else:
                    sheet_new.cell(idrow_new,2).value = "N/A"

                sheet_new.cell(idrow_new,3).value = TAG

                if filter > 0:
                    if sheet_old.cell(idrow_old, filter).value is not None:
                        sheet_new.cell(idrow_new,4).value = sheet_old.cell(idrow_old, filter).value
                    else:
                        sheet_new.cell(idrow_new,4).value = "N/A"
                else:
                    pass
                idrow_new += 1
                exitcode = 1
                cells +=1
            else:
                break
        idrow_old += 1
        idrow_new += 1



def endProgram():
    os.system("cls")
    time.sleep(0.5)

    if exitcode:
        print("                                                                ")
        print("                                                                ")
        print("                                                                ")
        o.save(filename=output)
        print("________________________________________________________________")
        print("|                                                              |")
        print("|                       SUCCESFUL EXPORT!                      | ")
        print(f"|                EXCEL FILE {output} CREATED                   |")
        print(f"|                  {cells} CELLS CREATED                      |")
        print("|                                                              |")
        print("|             PLEASE, PRESS ANY KEY TO EXIT...                |")
        print("|______________________________________________________________|")
        input("\n>> ")



#EXIT CODE SETUP
exitcode = 0
