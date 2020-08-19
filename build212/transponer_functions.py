

import consolemenu
from colorama import init
from colorama import Fore, Back, Style
import openpyxl
import os
import time


# DICTIONARY DECLARATION
columnDictionary = {
    "A" : 1,
    "B" : 2,
    "C" : 3,
    "D" : 4,
    "E" : 5,
    "F" : 6,
    "G" : 7,
    "H" : 8,
    "I" : 9,
    "J" : 10,
    "K" : 11,
    "L" : 12,
    "M" : 13,
    "N" : 14,
    "O" : 15,
    "P" : 16,
    "Q" : 17,
    "R" : 18,
    "S" : 19,
    "T" : 20,
    "U" : 21,
    "V" : 22,
    "W" : 23,
    "X" : 24,
    "Y" : 25,
    "Z" : 26

}

# COLOR FUNCTIONS

def outputcolor():
    print(Fore.WHITE + Back.RED + Style.BRIGHT)
def inputcolor():
    print(Fore.GREEN + Back.BLACK + Style.BRIGHT)
def resetandclose():
    print(Style.RESET_ALL)

#VARIBALE DEFINITION

menu = None
wd = None
selection = None

filename = None
file = None
o = None
output = None

sheet_old = None
sheet_new = None

scolumnname = None
idcolumn_old = None
idcolumn_new = None
startcolumn = None
endcolumn = None

idrow_old = None
idrow_new = 1
cells = 0
TAG = None


# ERROR HANDLING
def trycolumn():
    try:
        columnDictionary[scolumnname]
    except:
        outputcolor()
        print("                        ERROR:                            ")
        print("      LA COLUMNA NO ES VÁLIDA, VUELVA A INTENTARLO.       ")
        return False
    else:
        return True


def mainMenu():
    menu = consolemenu.SelectionMenu(["INICIAR"],title="TRASPONER v.2", subtitle="UTILIZE ESTA HERRAMIENTA PARA FÁCILMENTE TRASPONER CELDAS.",prologue_text="ESTE PROGRAMA USA LOS MÓDULOS OPENPYXL, COLORAMA, OS, TIME, Y CONSOLEMENU", epilogue_text="HECHO POR: FRANCISCO SALCEDO 2020 ", exit_option_text="SALIR")
    menu.show()
    menu.join()

def columnWarning():
    print("______________________________________________________________________________")
    print("|         A CONTINUACIÓN, INGRESARÁ LA LETRA DE LA COLUMNA A INDICAR.         |")
    print("|              POR FAVOR, INGRESE SOLO LA LETRA EN MAYÚSCULA                  |")
    print("|            EL PROGRAMA NO VERIFICARA SI LA COLUMNA ESTA VACÍA.              |")
    print("|           PARA ESTO, PUEDE ABRIR EL ARCHIVO EXCEL PARA VERIFICAR.           |")
    print("|               UTILIZE ESTA HERRAMIENTA BAJO SU PROPIO RIESGO                |")
    print("|                                                                             |")
    print("|                     PRESIONE LA TECLA ENTER PARA CONTINUAR                  |")
    print("|_____________________________________________________________________________|")
    input("\n>> ")
    os.system("cls")
# INPUT FUNCTIONS

def newquery():
    global exitcode
    options= ["DESEO EXTRAER MAS DE ESTE ARCHIVO", "TERMINAR CON ESTE ARCHIVO"]
    menu = consolemenu.SelectionMenu(options,title="DESEA CONTINUAR CON ESTE ARCHIVO?",subtitle="LA EXPORTACIÓN SE DARÁ EN UN SOLO ARCHIVO",epilogue_text=" ",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    if selection == 0:
        return True
    elif selection == 1:
        return False
def openFile():
    global wd
    global menu
    global selection
    global filename
    global file
    global o

    wd = os.listdir()
    print(wd)
    menu = consolemenu.SelectionMenu(wd,title="ESCOJA EL ARCHIVO DEL CUAL EXTRAER DATOS",subtitle="SOLO UTILIZE FORMATOS .XLSX O .XLS",epilogue_text="SI NO LOCALIZA SU ARCHIVO, POR FAVOR VERIFIQUE QUE ÉSTE EJECUTABLE ESTÉ EN EL MISMO DIRECTORIO.",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    filename = wd[selection]
    try:
        openpyxl.load_workbook(filename)
    except:
        print("                                 ERROR:                                   ")
        print("                    HA SELECCIONADO UN ARCHIVO INVÁLIDO.                  ")
        print("                             INTENTE DE NUEVO                             ")
        input(">> ")
        openFile()
    else:
        file = openpyxl.load_workbook(filename)
        o = openpyxl.Workbook()
        time.sleep(0.2)

def openSheet():
    global wd
    global menu
    global selection
    global sheet_old
    wd = file.sheetnames
    menu = consolemenu.SelectionMenu(wd,title="ESCOJA LA HOJA DE TRABAJO DE LA CUAL EXTRAER DATOS",subtitle="VERIFIQUE QUE SEA LA CORRECTA.",epilogue_text="PUEDE ABRIR EL ARCHIVO EXCEL PARA VERIFICAR.",show_exit_option=False)
    menu.show()
    menu.join()
    selection = menu.selected_option
    global fixdopt
    global fixddir
    fixdopt = menu.selected_option
    fixddir = file.sheetnames
    sheet_old = file.worksheets[file.sheetnames.index(wd[selection])]

    time.sleep(0.2)

def assignColumns():

    global scolumnname
    global idcolumn_old
    global startcolumn
    global endcolumn
    global TAG

    print("___________________________________________________________")
    print("|                                                         |")
    print("|         INGRESE LA COLUMNA DONDE SE ENCUENTRAN:         |")
    print("|                        SU KEY                           |")
    print("|                   EJ. SHOP ID, X...                     |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    idcolumn_old = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")


    print("___________________________________________________________")
    print("|                                                         |")
    print("|      INGRESE LA COLUMNA DONDE EMPIEZA SU RANGO          |")
    print("|                   EJ. ITEM, DATA...                     |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    startcolumn = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")

    print("___________________________________________________________")
    print("|                                                         |")
    print("|      INGRESE LA COLUMNA DONDE TERMINA SU RANGO          |")
    print("|                   EJ. ITEM, DATA...                     |")
    print("|_________________________________________________________|")
    scolumnname = input("\n>> ").capitalize()
    while not trycolumn():
        scolumnname = input(">> ").capitalize()
    endcolumn = columnDictionary[scolumnname]
    time.sleep(0.2)
    os.system("cls")

    print("___________________________________________________________")
    print("|                                                         |")
    print("|                                                         |")
    print("|        INGRESE UN TAG PARA IDENTIFICAR LA HOJA:         |")
    print(f"|                   {fixddir[fixdopt]}                    |")
    print("|_________________________________________________________|")
    TAG = input("\n>> ")
    time.sleep(0.2)
    os.system("cls")

def setExport():
    global sheet_new
    global output
    print("___________________________________________________________")
    print("|                                                         |")
    print("|       INGRESE EL NOMBRE DEL ARCHIVO A CREAR:            |")
    print("|                 ESTE DEBE SER NUEVO.                    |")
    print("|_________________________________________________________|")

    output = input("\n>> ")+".xlsx"
    sheet_new = o.worksheets[0]
    time.sleep(0.2)
    os.system("cls")


def compute():

    global idrow_new
    global idrow_old
    global exitcode
    global cells

    idrow_old = 1
    idcolumn_new = 1

    while sheet_old.cell(idrow_old, idcolumn_old).value is not None:
        for column in range(startcolumn,endcolumn+1):
            if sheet_old.cell(idrow_old, column).value is not None:
                sheet_new.cell(idrow_new, idcolumn_new).value = sheet_old.cell(idrow_old, idcolumn_old).value
                sheet_new.cell(idrow_new,2).value = sheet_old.cell(idrow_old, column).value
                sheet_new.cell(idrow_new,3).value = TAG
                idrow_new += 1
                exitcode = 1
                cells +=1
            else:
                break
        idrow_old += 1
        idrow_new += 1



def endProgram():
    os.system("cls")
    time.sleep(0.5)
    if exitcode:
        print("                                                                ")
        print("                                                                ")
        print("                                                                ")
        o.save(filename=output)
        print("________________________________________________________________")
        print("|                                                              |")
        print("|                     EXPORTACIÓN EXITOSA                      | ")
        print(f"|                  ARCHIVO {output} CREADO                     |")
        print("|    RECUERDE GUARDAR EL ARCHIVO EN EL FORMATO CORRECTO        |")
        print(f"|                  {cells} CELDAS AÑADIDAS                     |")
        print("|                                                              |")
        print("|             PRESIONE LA TECLA ENTER PARA SALIR               |")
        print("|______________________________________________________________|")
        input("\n>> ")



#EXIT CODE SETUP
exitcode = 0
